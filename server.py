"""
server.py — PaisaMap local server

  python3 server.py          → serves at http://localhost:8080
  python3 server.py --port 9090

Endpoints:
  /                          static → index.html
  /data/...                  static → data/ files (CSV, etc.)
  /api/reverse?lat=&lng=     Nominatim reverse geocode proxy (server IP, avoids browser rate-limit)
  /api/search?q=QUERY        Nominatim search proxy
  /api/enrich                GET {pincode, lat, lng, name, source} — enriches new pincode
  /api/status/<pincode>      enrichment job status
  /api/enrich_stats          enrichment log summary (counts by source, recent activity)
  /api/export                GET {format, dataset, scope, lat, lng, radius_km} — download
                              PPI+signals or the enrichment audit log as CSV/JSON/XLSX
  /api/db_status             DB dual-write status + row-count parity vs. the CSVs
                              (see paisamap-etl/etl/_db.py — no-op until DATABASE_URL is set)
  /api/auth/google           POST {credential} — verify Google ID token, create session
  /api/auth/logout           POST — clear session
  /api/auth/me               GET current user/plan/credits, 401 if not signed in
  /api/projects              GET/POST — CRUD for the signed-in user's projects
  /api/locations             GET/POST — saved locations (POST with no project_id
                              attaches to the auto-created "Saved Locations" project)
  /api/activity               GET — the signed-in user's recent activity feed
  /api/credits                GET — credit balance + ledger
  /api/reports                GET — reports list; POST {project_id} — generate a PDF
                              (economic score/benchmark/opportunity/suitability/risk per
                              saved location); GET /api/reports/<id>/download — fetch the PDF
  /api/intelligence/score      GET ?pincode=[&avg_ticket=&target_segment=&business_type=]
                              economic score, benchmarks, executive summary, risk, and
                              (if a business profile is given) opportunity/suitability
  /api/intelligence/compare    GET ?pincodes=A,B,C (1-8) — same, ranked, side by side
                              (see blueprints/ — 503 if DATABASE_URL isn't set, no CSV fallback;
                              intelligence/* is the exception — no DB required, same as /api/export)

REPORTS_DIR (env var, optional) — where generated report PDFs are written.
Defaults to ./data/reports for local dev. Production MUST point this outside
the git-tracked tree (see blueprints/reports.py's docstring) — deploy.sh
mirrors a fresh git checkout onto the nginx static root with `rsync --delete`
and would silently wipe any PDF that only ever existed in that destination copy.
"""

import csv
import io
import json
import os
import re
import secrets
import shutil
import sys
import threading
import subprocess
import urllib.request
import urllib.parse
from datetime import datetime, timezone, timedelta
from pathlib import Path
from flask import Flask, request, jsonify, send_file, send_from_directory

APP  = Path(__file__).parent
ETL  = APP / "paisamap-etl"
ENRICH_SCRIPT = ETL / "etl" / "enrich_single.py"
VENV_PY = ETL / "venv" / "bin" / "python3"
PYTHON  = str(VENV_PY) if VENV_PY.exists() else sys.executable

ETL_RAW = ETL / "data" / "raw"
ETL_OUT = ETL / "data" / "output"

# Optional DB dual-write (see paisamap-etl/etl/_db.py) — no-op unless
# DATABASE_URL is set, and unavailable entirely if sqlalchemy isn't
# installed in whatever interpreter is running this Flask app. Either way
# the CSV-based flow below is unaffected.
sys.path.insert(0, str(ETL / "etl"))
try:
    import _db
except ImportError:
    _db = None

ENRICH_LOG     = APP / "data" / "output" / "enrichment_log.csv"
PPI_MAP_DATA   = APP / "data" / "output" / "ppi_map_data.csv"
LOG_FIELDS     = ["timestamp", "pincode", "name", "lat", "lng", "source", "ppi", "income"]

# nginx serves the live map's /data/... from a separate static mirror
# (/var/www/paisamap), refreshed only by a full deploy — so a pincode
# enriched on demand here doesn't reach the map until the next deploy
# unless we also copy it there directly. No-op if the mirror doesn't exist
# (e.g. local dev via `python3 server.py`).
STATIC_MIRROR_OUT = Path("/var/www/paisamap/data/output")


def _mirror_to_static():
    if not STATIC_MIRROR_OUT.is_dir():
        return
    for f in (ENRICH_LOG, PPI_MAP_DATA):
        if f.exists():
            shutil.copy2(f, STATIC_MIRROR_OUT / f.name)


# ── Export: PPI/income/spend joined with every pincode-level raw signal ────────
# Moved to paisamap-etl/etl/_signals_data.py (Phase 2) so blueprints/intelligence.py
# can reuse the exact same join without importing server.py back into a blueprint
# it registers. Pull the names in locally so every existing reference below
# (EXPORT_CORE_FIELDS, EXPORT_ALL_COLUMNS, etc.) keeps working unchanged.
from _signals_data import (EXPORT_CORE_FIELDS, EXPORT_ALL_COLUMNS,
                            haversine_km as _haversine_km, coerce as _coerce,
                            load_ppi_signals_rows as _load_ppi_signals_rows)

app = Flask(__name__)

# Session config for the auth blueprints below. SECRET_KEY should be set in
# production (injected via /etc/paisamap/db.env, same EnvironmentFile= pattern
# as DATABASE_URL) — an ephemeral random key is fine for local dev but means
# every restart invalidates existing sessions, so it's not acceptable in prod.
app.secret_key = os.environ.get("SECRET_KEY") or secrets.token_hex(32)
_db_is_postgres = os.environ.get("DATABASE_URL", "").startswith("postgresql")
app.config.update(
    SESSION_COOKIE_NAME="pm_session",
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SAMESITE="Lax",
    SESSION_COOKIE_SECURE=_db_is_postgres,  # reuses the sqlite(dev)/postgresql(prod)
                                             # convention _db.py already established,
                                             # instead of a new env var
    PERMANENT_SESSION_LIFETIME=timedelta(days=30),
)

try:
    from blueprints.auth import auth_bp
    from blueprints.projects import projects_bp
    from blueprints.locations import locations_bp
    from blueprints.activity import activity_bp
    from blueprints.credits import credits_bp
    from blueprints.reports import reports_bp
    from blueprints.intelligence import intelligence_bp
    app.register_blueprint(auth_bp)
    app.register_blueprint(projects_bp)
    app.register_blueprint(locations_bp)
    app.register_blueprint(activity_bp)
    app.register_blueprint(credits_bp)
    app.register_blueprint(reports_bp)
    app.register_blueprint(intelligence_bp)
except ImportError as e:
    print(f"[server] auth/workspace blueprints unavailable: {e}", flush=True)

# Job registry: pincode → {status, ppi, log, error, source}
_jobs: dict = {}
_lock = threading.Lock()


# ── Enrichment log ────────────────────────────────────────────────────────────
def _append_log(pc, name, lat, lng, source, ppi, income):
    row = {
        "timestamp": datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
        "pincode":   pc,
        "name":      name,
        "lat":       lat,
        "lng":       lng,
        "source":    source,   # phase1 | yah | prefetch | search | manual
        "ppi":       ppi,
        "income":    income,
    }
    need_header = not ENRICH_LOG.exists() or ENRICH_LOG.stat().st_size == 0
    ENRICH_LOG.parent.mkdir(parents=True, exist_ok=True)
    with open(ENRICH_LOG, "a", newline="") as f:
        w = csv.DictWriter(f, fieldnames=LOG_FIELDS)
        if need_header:
            w.writeheader()
        w.writerow(row)

    # Dual-write to the database (no-op unless DATABASE_URL is set). CSV
    # above is already durable — a DB hiccup here must never surface to
    # the /api/enrich caller, whose job already succeeded.
    if _db is not None:
        try:
            _db.insert_log(row["timestamp"], pc, name, lat, lng, source, ppi, income)
        except Exception as e:
            print(f"  WARN: DB log dual-write failed: {e}", flush=True)


def _parse_enrich_output(stdout):
    """Extract PPI and income from enrich_single.py stdout."""
    ppi    = re.search(r"PPI \(ML\):\s*(\d+)",           stdout)
    income = re.search(r"household income:\s*₹([\d,]+)", stdout)
    return (
        ppi.group(1)    if ppi    else "",
        income.group(1) if income else "",
    )


# ── Static files ──────────────────────────────────────────────────────────────
@app.route("/")
def root():
    return send_file(APP / "index.html")

@app.route("/data/<path:fname>")
def serve_data(fname):
    return send_from_directory(APP / "data", fname)

@app.route("/assets/<path:fname>")
def serve_assets(fname):
    # nginx serves the whole app directory statically in prod, so assets/ (logo,
    # favicon) resolve there with no extra config — this dev server only had
    # explicit routes for "/" and "/data/", so the same paths 404'd locally
    # (found via the workspace app's sidenav logo failing to load in local testing).
    return send_from_directory(APP / "assets", fname)


# ── Nominatim proxies ─────────────────────────────────────────────────────────
@app.route("/api/reverse")
def api_reverse():
    """Proxy Nominatim reverse geocode — server IP avoids browser rate-limit."""
    lat = request.args.get("lat", "").strip()
    lng = request.args.get("lng", request.args.get("lon", "")).strip()
    if not lat or not lng:
        return jsonify({"error": "lat and lng required"}), 400
    try:
        params = urllib.parse.urlencode({
            "lat": lat, "lon": lng, "zoom": 14,
            "format": "json", "addressdetails": 1,
        })
        url = f"https://nominatim.openstreetmap.org/reverse?{params}"
        req = urllib.request.Request(
            url, headers={"User-Agent": "PaisaMap-Server/1.0",
                          "Accept-Language": "en",
                          "Accept": "application/json"})
        with urllib.request.urlopen(req, timeout=10) as r:
            return r.read(), 200, {"Content-Type": "application/json",
                                   "Access-Control-Allow-Origin": "*"}
    except Exception as e:
        return jsonify({"error": str(e)}), 502


@app.route("/api/search")
def api_search():
    q = request.args.get("q", "").strip()
    if not q:
        return jsonify([])
    try:
        params = urllib.parse.urlencode({
            "format": "jsonv2", "limit": 6,
            "polygon_geojson": 1, "addressdetails": 1,
            "countrycodes": "in", "q": q
        })
        url = f"https://nominatim.openstreetmap.org/search?{params}"
        req = urllib.request.Request(
            url, headers={"User-Agent": "PaisaMap-Server/1.0",
                          "Accept": "application/json"})
        with urllib.request.urlopen(req, timeout=10) as r:
            return r.read(), 200, {"Content-Type": "application/json",
                                   "Access-Control-Allow-Origin": "*"}
    except Exception as e:
        return jsonify({"error": str(e)}), 502


# ── Enrichment ────────────────────────────────────────────────────────────────
@app.route("/api/enrich")
def api_enrich():
    pc     = request.args.get("pincode", "").strip()
    lat    = request.args.get("lat",     "").strip()
    lng    = request.args.get("lng",     "").strip()
    name   = request.args.get("name",    pc).strip()
    source = request.args.get("source",  "yah").strip()   # yah | prefetch | search | manual

    if not pc or not lat or not lng:
        return jsonify({"error": "pincode, lat, lng are required"}), 400

    with _lock:
        job = _jobs.get(pc, {})
        if job.get("status") in ("running", "done"):
            return jsonify(job)
        _jobs[pc] = {"status": "running", "source": source}

    threading.Thread(target=_run_enrich, args=(pc, lat, lng, name, source),
                     daemon=True).start()
    return jsonify({"status": "started", "source": source})


@app.route("/api/status/<pincode>")
def api_status(pincode):
    with _lock:
        return jsonify(_jobs.get(pincode, {"status": "unknown"}))


@app.route("/api/health")
def health():
    return jsonify({"status": "ok"})


@app.route("/api/config")
def api_config():
    """Public, non-secret runtime config for frontends — Google OAuth client IDs
    are meant to be visible client-side (unlike SECRET_KEY/DATABASE_URL). Lets
    the /workspace React app pick up the same client ID that index.html gets via
    deploy.sh's sed injection, without a second injection mechanism."""
    return jsonify({"google_client_id": os.environ.get("GOOGLE_CLIENT_ID", "")})


@app.route("/api/db_status")
def db_status():
    """Whether the DB dual-write is configured, and what it currently holds
    — CSV row counts alongside it as a parity check."""
    if _db is None:
        return jsonify({"enabled": False, "reason": "sqlalchemy not importable"})
    if not _db.enabled():
        return jsonify({"enabled": False, "reason": "DATABASE_URL not set"})
    try:
        db_counts = _db.counts()
    except Exception as e:
        return jsonify({"enabled": True, "error": str(e)}), 500

    csv_pincodes = 0
    ml_path = ETL_OUT / "ppi_ml_refined.csv"
    if ml_path.exists():
        with open(ml_path, newline="") as f:
            csv_pincodes = sum(1 for _ in csv.DictReader(f))
    csv_log = 0
    if ENRICH_LOG.exists():
        with open(ENRICH_LOG, newline="") as f:
            csv_log = sum(1 for _ in csv.DictReader(f))

    return jsonify({
        "enabled": True,
        "db":  db_counts,
        "csv": {"pincodes": csv_pincodes, "enrichment_log": csv_log},
    })


@app.route("/api/enrich_stats")
def enrich_stats():
    """Return enrichment counts by source and the last 20 enriched pincodes."""
    if not ENRICH_LOG.exists():
        return jsonify({"total": 0, "by_source": {}, "recent": []})

    rows = []
    try:
        with open(ENRICH_LOG, newline="") as f:
            rows = list(csv.DictReader(f))
    except Exception:
        pass

    by_source: dict = {}
    for r in rows:
        src = r.get("source", "unknown")
        by_source[src] = by_source.get(src, 0) + 1

    recent = [
        {"pincode": r["pincode"], "name": r["name"], "source": r["source"],
         "ppi": r["ppi"], "ts": r["timestamp"]}
        for r in rows[-20:][::-1]
    ]

    return jsonify({
        "total":     len(rows),
        "by_source": by_source,
        "recent":    recent,
    })


# ── Export ────────────────────────────────────────────────────────────────────
# _haversine_km / _coerce / _load_ppi_signals_rows now live in _signals_data.py
# (imported above) — kept as module-level names here via `as` aliases so every
# call site below is unchanged.


def _load_log_rows():
    """Returns (rows, source_label) — DB if configured, else enrichment_log.csv."""
    if _db is not None:
        db_rows = _db.fetch_log()
        if db_rows is not None:
            return db_rows, "database"
    if not ENRICH_LOG.exists():
        return [], "csv"
    with open(ENRICH_LOG, newline="") as f:
        return list(csv.DictReader(f)), "csv"


@app.route("/api/export")
def api_export():
    fmt     = request.args.get("format", "csv").lower()
    dataset = request.args.get("dataset", "ppi").lower()
    scope   = request.args.get("scope", "all").lower()
    lat     = request.args.get("lat", type=float)
    lng     = request.args.get("lng", type=float)
    radius  = request.args.get("radius_km", type=float)
    columns_param = request.args.get("columns", "").strip()

    if fmt not in ("csv", "json", "xlsx"):
        return jsonify({"error": "format must be csv, json, or xlsx"}), 400
    if dataset not in ("ppi", "log"):
        return jsonify({"error": "dataset must be ppi or log"}), 400

    ts = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")

    if dataset == "log":
        rows, source = _load_log_rows()
        columns    = LOG_FIELDS
        sheet_name = "Enrichment Log"
        base_name  = "paisamap_enrichment_log"
    else:
        joined, source = _load_ppi_signals_rows()
        if scope == "view" and lat is not None and lng is not None and radius:
            joined = {
                pc: r for pc, r in joined.items()
                if r.get("lat") and r.get("lng")
                and _haversine_km(lat, lng, float(r["lat"]), float(r["lng"])) <= radius
            }
        rows       = sorted(joined.values(),
                             key=lambda r: _coerce(r.get("ppi_ml")) or 0, reverse=True)
        columns    = EXPORT_ALL_COLUMNS
        if columns_param:
            requested = set(columns_param.split(","))
            # core identity/PPI fields always ride along; signal columns are opt-in
            columns = [c for c in EXPORT_ALL_COLUMNS if c in EXPORT_CORE_FIELDS or c in requested]
        sheet_name = "PPI & Signals"
        base_name  = "paisamap_ppi_signals"

    filename = f"{base_name}_{ts}.{fmt}"

    if fmt == "json":
        payload = json.dumps({
            "generated_at": ts,
            "dataset":      dataset,
            "source":       source,
            "count":        len(rows),
            "rows":         [{k: _coerce(v) for k, v in r.items() if k in columns} for r in rows],
        }, indent=2)
        return payload, 200, {
            "Content-Type":        "application/json",
            "Content-Disposition": f"attachment; filename={filename}",
            "X-Data-Source":       source,
        }

    if fmt == "csv":
        buf = io.StringIO()
        w = csv.DictWriter(buf, fieldnames=columns, extrasaction="ignore")
        w.writeheader()
        for r in rows:
            w.writerow(r)
        return buf.getvalue(), 200, {
            "Content-Type":        "text/csv",
            "Content-Disposition": f"attachment; filename={filename}",
            "X-Data-Source":       source,
        }

    # xlsx
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font
    except ImportError:
        return jsonify({"error": "Excel export unavailable — openpyxl not installed on server"}), 501

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(columns)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    ws.freeze_panes = "A2"
    for r in rows:
        ws.append([_coerce(r.get(c, "")) for c in columns])
    for i, c in enumerate(columns, 1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = max(10, min(24, len(c) + 2))

    meta = wb.create_sheet("Metadata")
    meta.append(["Generated at (UTC)", ts])
    meta.append(["Dataset", sheet_name])
    meta.append(["Row count", len(rows)])
    meta.append(["Data read from", "PostgreSQL" if source == "database" else "CSV"])
    meta.append(["About", "PaisaMap — pan-India PPI model + government/open data signals"])
    meta.append(["Note", "Modelled estimates, not real transaction records — see in-app disclaimer."])

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out.getvalue(), 200, {
        "Content-Type":        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "Content-Disposition": f"attachment; filename={filename}",
        "X-Data-Source":       source,
    }


# ── Worker ────────────────────────────────────────────────────────────────────
def _run_enrich(pc, lat, lng, name, source="yah"):
    try:
        res = subprocess.run(
            [PYTHON, str(ENRICH_SCRIPT), pc, lat, lng, name],
            capture_output=True, text=True, timeout=180, cwd=str(ETL)
        )
        ppi, income = _parse_enrich_output(res.stdout)
        with _lock:
            if res.returncode == 0:
                lines = [l for l in res.stdout.splitlines() if l.strip()]
                ppi_line = next((l for l in reversed(lines) if "PPI" in l), "")
                _jobs[pc] = {
                    "status":  "done",
                    "source":  source,
                    "ppi":     ppi,
                    "income":  income,
                    "log":     res.stdout[-3000:],
                    "summary": ppi_line,
                }
                _append_log(pc, name, lat, lng, source, ppi, income)
                _mirror_to_static()
            else:
                _jobs[pc] = {"status": "error", "source": source,
                             "error": res.stderr[-1500:]}
    except Exception as e:
        with _lock:
            _jobs[pc] = {"status": "error", "source": source, "error": str(e)}


# ── Entry point ───────────────────────────────────────────────────────────────
if __name__ == "__main__":
    port = 8080
    for i, a in enumerate(sys.argv[1:]):
        if a == "--port" and i + 2 < len(sys.argv):
            port = int(sys.argv[i + 2])

    print(f"\n  PaisaMap server  →  http://localhost:{port}")
    print(f"  App:             {APP}")
    print(f"  ETL:             {ETL}")
    print(f"  Python:          {PYTHON}\n")
    app.run(port=port, debug=False, use_reloader=False)
